import os
import re
import logging
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List

import openpyxl
from django.apps import apps
from django.conf import settings
from django.core.management.base import BaseCommand
import deep_translator

logger = logging.getLogger(__name__)


@dataclass
class DatabaseExtractor:
    module_name: str
    attributes: list
    filters: dict = field(default_factory=dict)
    sub_extractors: dict = field(default_factory=dict)
    unique: bool = True
    auto_translate: bool = False

    @property
    def model_name(self):
        return self.module_name.split(".")[-1]

    @property
    def model(self):
        return apps.get_model(*self.module_name.split("."))
    
    @property
    def queryset(self):
        return self.model.objects.filter(**self.filters).prefetch_related(*self.sub_extractors.keys()).order_by("pk")


def ext(module_name, attributes, **kwargs):
    return DatabaseExtractor(module_name=module_name, attributes=attributes, **kwargs)


STATUS_PUBLISHED = 3
SEPARATOR = ";"
DATABASE_EXTRACTORS = [
    ext("recipe_mgr.Recipe", ("name", "summary"), unique=False, filters={"status": STATUS_PUBLISHED, "internal": False}, auto_translate=True,
        sub_extractors={"instructions": ext("recipe_mgr.RecipeInstruction", ("text", ), unique=False, auto_translate=True)}
    ),
    ext("recipe_mgr.RecipeTag", ("name", ), auto_translate=False),
    ext("recipe_mgr.UstensilCategory", ("name", ), auto_translate=False),
    ext("recipe_mgr.Ustensil", ("name", ), auto_translate=False),
    ext("recipe_mgr.DishType", ("name", ), auto_translate=False),
    ext("recipe_mgr.CookingMethod", ("name", ), auto_translate=False),
    ext("recipe_mgr.Food", ("name", ), filters={"enabled": True}, auto_translate=False,  # full_name"
        sub_extractors={"conversions": ext("recipe_mgr.FoodConversion", ("unit", "unit_plural"), auto_translate=False)}
    ),
    ext("recipe_mgr.RawState", ("name", ), auto_translate=False),
    ext("recipe_mgr.FoodTag", ("name", ), auto_translate=False),
    ext("recipe_mgr.FoodType", ("name", ), auto_translate=False),
    ext("shopping_mgr.ShoppingCategory", ("name", ), auto_translate=False),
    ext("user_mgr.ConfigStage", ("name", "express_description", "description"), auto_translate=False),
    # ext("user_mgr.Role", ("name", "description")),
    ext("diet_mgr.Diet", ("title", "description"), filters={"enabled": True}, auto_translate=False,  # TODO: specify slim ?
        # sub_extractors={"parameters": ext("diet_mgr.UserDietParameter", ("name", ))}
    ),
    ext("location_mgr.Location", ("name", ), auto_translate=False),
    ext("nutrient.Nutrient", ("short_name", "unit"), filters={"enabled": True}, auto_translate=False),  # "name",
    # ext("nutrient.NutrientPack", ("title", "description", "warning")),
    ext("planning_mgr.MealType", ("name", "nickname"), auto_translate=False),
    ext("profile_mgr.ProfileMetric", ("name", "unit", "description"), auto_translate=False),
]


def find_code(codes, supported_languages, translator_name):
    for code in codes:
        if code in supported_languages:
            return code
    raise Exception(f"Target language '{codes}' not supported by translator '{translator_name}'\nSupported languages are:\n - " + "\n - ".join(supported_languages))


def generate_translator(translator_name, source_codes, target_codes):
    print(f" + Registering translator '{translator_name}'")
    cls = getattr(deep_translator, translator_name)
    supported_languages = cls.get_supported_languages()
    source = find_code(source_codes, supported_languages, translator_name)
    target = find_code(target_codes, supported_languages, translator_name)
    # proxies_example = {
    #     "https": "34.195.196.27:8080",
    #     "http": "34.195.196.27:8080"
    # }
    return cls(source=source, target=target)  # , proxies=proxies_example


TRANSLATORS = [generate_translator(translator_name, source_codes=("fr", "french"), target_codes=("italian", "it"))
    for translator_name in (
        "GoogleTranslator",
        # "MicrosoftTranslator",  # API key needed
        "PonsTranslator",
        # "LingueeTranslator",  # Nul
        # "MyMemoryTranslator",  # Nul
        # "YandexTranslator",  # API key needed
        # "PapagoTranslator",  # Removed ?
    )
]


@dataclass
class Wording:
    french: str = ""
    italian: str = ""
    italian_auto: dict = field(default_factory=dict)
    origins: set = field(default_factory=set)

    def translate(self, auto_translate=False):
        if not self.italian:
            self.italian = re.sub("[a-zéèàùçêïA-Z]", "*", self.french)
        if not auto_translate:
            return
        for translator in TRANSLATORS:
            try:
                self.italian_auto[translator.__class__.__name__] = translator.translate(self.french)
            except:
                self.italian_auto[translator.__class__.__name__] = "-"


IMMUTABILITY_CHECK = defaultdict(Wording)
ORDERED_WORDINGS = defaultdict(Wording)
UNIQUE_WORDINGS = defaultdict(Wording)
SHEET_NAME = "translations"
HEADER_ROW_NB = 1
EXCEL_MAX_CELL_LENGTH = 32767


def ensure_immutability(origin, text):
    if origin in IMMUTABILITY_CHECK:
        if IMMUTABILITY_CHECK[origin] != text:
            raise Exception(f"French text has changed in database '{origin}': {IMMUTABILITY_CHECK[origin]} --> {text}")
    else:
        IMMUTABILITY_CHECK[origin] = text


def get_module_name(instance):
    return f"{instance._meta.app_label}.{instance._meta.object_name}"


def batch(iterable, n=1):
    iterable = list(iterable)
    l = len(iterable)
    for i in range(0, l, n):
        yield iterable[i:min(i + n, l)]


class Command(BaseCommand):
    help = 'Extract translations into an excel file'

    LETTER = re.compile("[^a-zA-Z]")

    def add_arguments(self, parser):
        parser.add_argument('action', type=str)
        parser.add_argument('-o', '--output', type=str, default="translations.xlsx")
        parser.add_argument('-i', '--input', type=str, default="translations.xlsx")
        parser.add_argument('-l', '--language', type=str, default="it")

    def clean_string(self, value):
        value = value or ""
        return value.strip()

    def add_string(self, module, pk, attr, text, extractor):
        print(".", end='')
        text = self.clean_string(text)
        origin = f"{module}:{pk}:{attr}"
        ensure_immutability(origin, text)
        if extractor.unique:
            wording = UNIQUE_WORDINGS[text]
        else:
            wording = ORDERED_WORDINGS[origin]
        wording.french = text
        wording.origins.add(origin)
        # Ultimately, translations should be done only if needed
        if len(wording.italian_auto) == len(self.translator_columns) and None not in wording.italian_auto.values():
            return
        wording.translate(extractor.auto_translate)

    def write_row(self, wording):
        self.character_count += len(re.sub(self.LETTER, "", wording.french))
        self.word_count += len(wording.french.split(" "))
        self.sentences_count += 1
        for origins in batch(wording.origins, 500):
            origins = "\n".join(wording.origins)
            self.sheet.cell(column=1, row=self.row, value=origins)
            self.sheet.cell(column=2, row=self.row, value=wording.french)
            self.sheet.cell(column=3, row=self.row, value=wording.italian)
            for translator_name, italian in wording.italian_auto.items():
                if translator_name not in self.translator_columns:
                    # This translator seems to have been removed
                    logger.error(f"Translator removed {translator_name}")
                    continue
                self.sheet.cell(column=self.translator_columns[translator_name], row=self.row, value=italian)
            self.row += 1

    def write_translations(self, output):
        self.translator_columns = dict((translator.__class__.__name__, i + 4) for i, translator in enumerate(TRANSLATORS))
        workbook = openpyxl.Workbook()
        self.sheet = workbook.active
        self.sheet.name = SHEET_NAME
        # Header
        self.row = HEADER_ROW_NB
        headers = Wording("French", "Italian", origins=set(["origin"]), italian_auto=dict((c, c) for c in self.translator_columns.keys()))
        self.write_row(headers)
        # Content
        for wording in ORDERED_WORDINGS.values():
            self.write_row(wording)
        for wording in UNIQUE_WORDINGS.values():
            self.write_row(wording)
        workbook.save(output)
    
    def extract_instance(self, obj, extractor):
        module = get_module_name(obj)
        for attr in extractor.attributes:
            value = getattr(obj, attr, None) or ""
            value = value.strip()
            if not value:
                continue
            if type(value) is not str:
                raise Exception(f"Invalid data type {type(value)} on {module}:{obj.pk}:{attr}")
            self.add_string(module, obj.pk, attr, value, extractor)

    def extract_model(self, extractor):
        print(f" * Extracting model '{extractor.model_name}' ({extractor.queryset.count()} rows)")
        for relation, sub_extractor in extractor.sub_extractors.items():
            print(f"   - Extracting sub-model '{sub_extractor.model_name}' ({sub_extractor.queryset.count()} rows)")
        for instance in extractor.queryset:
            self.extract_instance(instance, extractor)
            if len(extractor.sub_extractors):
                for relation, sub_extractor in extractor.sub_extractors.items():
                    sub_queryset = getattr(instance, relation)
                    for sub_instance in sub_queryset.all().order_by("pk"):
                        self.extract_instance(sub_instance, sub_extractor)
        print("", flush=True)

    def action_extract(self):
        extension = ".xlsx"
        output = self.options["output"]
        if not output.endswith(extension):
            output += extension
        # Load existing translations
        input = self.options["input"]
        if input:
            self.load_translations(input)
        # Statistics
        self.character_count = 0
        self.word_count = 0
        self.sentences_count = 0
        # Extraction
        for extractor in DATABASE_EXTRACTORS:
            self.extract_model(extractor)
        self.write_translations(output)
        print(f"\nStatistics:\n - Strings: {self.sentences_count}\n - Words: {self.word_count}\n - Characters: {self.character_count}")

    def load_translations(self, input):
        if not input or not os.path.exists(input):
            raise Exception(f"Invalid input file '{input}'")
        workbook = openpyxl.load_workbook(filename=input)
        sheet = workbook.active
        # Headers
        self.translator_columns = dict(
            (sheet[chr(ord('A') + c) + "1"].value, chr(ord('A') + c))
            for c in range(3, sheet.max_column)
        )
        # Content
        for row in range(HEADER_ROW_NB + 1, sheet.max_row + 1):
            origins = set(sheet[f"A{row}"].value.split("\n"))
            french = sheet[f"B{row}"].value
            italian = sheet[f"C{row}"].value
            italian_auto = dict(
                (translator_name, sheet[f"{column}{row}"].value)
                for translator_name, column in self.translator_columns.items()
            )
            wording = Wording(french=french, italian=italian, origins=origins, italian_auto=italian_auto)
            first_origin = next(iter(origins))
            if first_origin.startswith("recipe_mgr.Recipe:") or first_origin.startswith("recipe_mgr.RecipeInstruction:"):
                assert len(origins) == 1
                ORDERED_WORDINGS[first_origin] = wording
            else:
                UNIQUE_WORDINGS[french] = wording
            for origin in origins:
                ensure_immutability(origin, french)

    def action_import(self):
        input = self.options["input"]
        language = self.options["language"]
        self.load_translations(input)
        statistics = defaultdict(int)
        self.import_wordings(ORDERED_WORDINGS.values(), language, statistics)
        self.import_wordings(UNIQUE_WORDINGS.values(), language, statistics)
        updates = "\n - ".join(f"{k}: {v}" for k, v in statistics.items())
        print(f"\nStatistics:\n - {updates}")
    
    def import_wordings(self, wordings, language, statistics):
        for wording in wordings:
            for origin in wording.origins:
                module_name, pk, attr = origin.split(":")
                model = apps.get_model(*module_name.split("."))
                value = {attr: wording.italian if language == "it" else wording.french}
                # model.objects.filter(pk=pk).update(**value)
                statistics[model.__name__] += 1

    def handle(self, **options):
        self.options = options
        action = options["action"]
        if not action:
            raise Exception("Missing action parameter")
        func = getattr(self, f"action_{action}", None)
        if not func:
            raise Exception(f"Invalid action '{action}'")
        func()
